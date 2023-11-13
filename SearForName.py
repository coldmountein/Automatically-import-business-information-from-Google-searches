import openpyxl
from DrissionPage import WebPage
from DrissionPage.common import ActionChains
from bs4 import BeautifulSoup
from DrissionPage import ChromiumPage
import math
import tkinter as tk
from tkinter import simpledialog
from DrissionPage import WebPage, ChromiumOptions, SessionOptions

def SaveName():
    # 选择搜索内容
    # 创建主窗口
    root = tk.Tk()
    root.title("1")  # 设置窗口标题
    # 弹出对话框获取用户输入
    SearchKeyWord = "パートナー募集 IT"
    SearchKeyWord = simpledialog.askstring("捜索キーワードを入力してください。", "捜索キーワードを入力してください。:")

    # 打印用户输入的内容
    print("キーワード:", SearchKeyWord)

    # 关闭主窗口
    root.destroy()

            # 选择搜索数量
        # 创建主窗口
    root2 = tk.Tk()
    root2.title("2")  # 设置窗口标题

    while True:
        # 弹出对话框获取用户输入
        try:
            SearchCount = int(simpledialog.askstring("捜索数を入力してください。", "捜索数を入力してください。:"))
            break
        # 打印用户输入的内容
        except ValueError:
        # 关闭主窗口]
            print("无效的输入，请重新输入。")        
    print("数:", SearchCount)
    root2.destroy()

    # 创建一个新的 Excel 工作簿
    workbook = openpyxl.Workbook()

    # 选择默认的工作表
    sheet = workbook.active
    # 创建页面对象，并启动或接管浏览器
    do = ChromiumOptions(read_file=False)  # 不读取文件方式新建配置对象
    so = SessionOptions(read_file=False)
    do.set_paths(browser_path=r'.\chrome.exe')  # 输入配置信息
    page = WebPage(driver_or_options=do, session_or_options=so)
    ac = ActionChains(page)
    # 跳转到登录页面
    page.get(f'https://www.google.com/search?q={SearchKeyWord}')

    flag = 0

    # 滚动到最下方
    while flag==0:
        ele = page.ele('#ofr')

        # 判断是否找到元素
        if ele:
            print('スクロール終わり')
            flag=1
            break
        if not ele:
            print('スクロール中')
            page.scroll.to_bottom()
            ele2 = page('.GNJvt ipz2Oe')
            if ele2:
                ele2.click()
                # print('找到了按钮')
            if not ele2:
                page.scroll.to_bottom()
                # print('没有没有找到了按钮')
    # 选择所有公司的名字
    ele3 = page.eles('.VuuXrf')
    ele4 = page.eles('@jsname=UWckNb')
    for index, value in enumerate(ele3[0:SearchCount * 2], start=1):
        if index % 2 == 1:
            print(value.inner_html)
            sheet.cell(row=math.ceil((index+1)/2), column=1, value=math.ceil(index/2))
            sheet.cell(row=math.ceil((index+1)/2), column=2, value=value.inner_html)
    for index, value in enumerate(ele4[0:SearchCount]):
        sheet.cell(row=index+1, column=7, value=value.link)
    # 选择所有公司的网址
    # print(page.ele('#card-section'))

    # page.scroll.to_bottom()
    # 保存 Excel 文件
    workbook.save("CompanyName.xlsx")