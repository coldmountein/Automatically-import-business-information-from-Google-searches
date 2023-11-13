import openpyxl
import requests
from bs4 import BeautifulSoup


def SaveUrl():
# 打开Excel文件
    workbook = openpyxl.load_workbook('CompanyName.xlsx')  # 替换为您的Excel文件路径

    # 选择要读取的工作表
    worksheet = workbook['Sheet']  # 替换为实际的工作表名称

    # 选择要读取的列号（以字母表示，例如'A'为第一列）
    column_letter = 'B'  # 替换为要读取的列字母

    # 读取特定列的所有数据并保存到数组
    column_data = []
    for cell in worksheet[column_letter]:
        column_data.append(cell.value)
    column_url = []


    # 打印数组中的数据（或根据需求执行其他操作）
    # print(column_data[0])
    # print(column_data)

    # for element in column_data:
    for index, element in enumerate(column_data, start=1):
        try:
            response = requests.get('https://doda.jp/DodaFront/View/CompanySearch.action', params={"q": {element},"sid":"CompanyTop_kw"})

            if response.status_code == 200:
                # 获取响应内容
                data = response.content

                # # 指定保存的文件路径和名称，例如以URL中的一部分作为文件名
                # file_name = {element}
                # file_path = f'output/{file_name}.html'  # 替换为您要保存的文件路径

                # # 写入数据到文件
                # with open(file_path, 'wb') as file:
                #     file.write(data)
                soup = BeautifulSoup(response.text, 'html.parser')
                # 要查找文本为"链接2"的所有<a>标签
                specific_links = soup.find('h2', class_='title')
                if (specific_links):
                    a_element = specific_links.find('a')
                    link= a_element['href']
                    column_url.append(f'{link}')
                    cell = worksheet.cell(row=index, column=11, value=f'{link}')
                    # print(f'{element}{link}')
                else:
                    print(f'{element}はdodaに登録していない')
                    column_url.append(f'{element}はdodaに登録していない')
                    cell = worksheet.cell(row=index, column=11, value='dodaに登録していない')

            else:
                print(f'Failed to fetch: {{element}}')
        except requests.exceptions.RequestException as e:
            print(f'Error while fetching {{element}}: {e}')
    # 保存 Excel 文件
    workbook.save('url.xlsx')

    # 关闭 Excel 文件
    workbook.close()