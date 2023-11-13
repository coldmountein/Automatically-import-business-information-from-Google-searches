import os
import SearchForUrl
import openpyxl
import requests
from bs4 import BeautifulSoup
import SearForName

SearForName.SaveName()
SearchForUrl.SaveUrl()
# 打开Excel文件
workbook = openpyxl.load_workbook('url.xlsx')  # 替换为您的Excel文件路径
    # 选择要读取的工作表
worksheet = workbook['Sheet']  # 替换为实际的工作表名称

    # 选择要读取的列号（以字母表示，例如'A'为第一列）
column_letter = 'K'  # 替换为要读取的列字母

    # 读取特定列的所有数据并保存到数组
column_data = []
for cell in worksheet[column_letter]:
    column_data.append(cell.value)
column_url = []

    # for element in column_data:
for index, element in enumerate(column_data, start=1):
    try:
        response = requests.get(element)

        if response.status_code == 200:
            # 获取响应内容
            data = response.content

            # # 指定保存的文件路径和名称，例如以URL中的一部分作为文件名
            # file_name = element.split("/")[-1]
            # file_path = f'output/{file_name}.html'  # 替换为您要保存的文件路径

            # # 写入数据到文件
            # with open(file_path, 'wb') as file:
            #     file.write(data)

            # print(f'Data from {element} saved to {file_path}')

            soup = BeautifulSoup(response.text, 'html.parser')
            # 要查找设立时间
            specific_links = soup.find('dl', class_='dateSet')
            settime = specific_links.find_all('span')
            # 查找在职人数
            specific_Person = soup.select('dl.dateSet')[1]
            PersonCount = specific_Person.select('span')
            # 查找注册资本金
            p_tag = soup.find("p", text=lambda text: text and "円" in text)
            Money = p_tag.text
            # 找事业内容
            
            # try:
            #     th1 = soup.find("th", text=lambda text: text and "事業内容" in text)                
            #     # td1 = th1.find_next_sibling('td')
            #     print(f'{element}')

            # except:
            #     # event_tag = td1.find('p')
            #     # event = event_tag.text
            #     event = 'No'
            # else:
            #     td1 = th1.find_next_sibling('td')
            #     event_tag = td1.find('p')
            #     event = event_tag.text

            if soup.find("th", text=lambda text: text and "事業内容" in text):
                th1 = soup.find("th", text=lambda text: text and "事業内容" in text)
                td1 = th1.find_next_sibling('td')
                event_tag = td1.find('p')
                event = event_tag.text
            else :
                event = 'dodaに入力していない'


            # 查找地区
            placelinks = soup.find('div', class_='address')
            place = placelinks.find_all('span')


            if (specific_links):
                # a_element = specific_links.find('a')
                # link= a_element['href']
                # column_url.append(f'{link}')
                # cell = worksheet.cell(row=index, column=11, value=f'{link}')
                # 打印找到的所有<th>标签
                
                cell = worksheet.cell(row=index, column=3, value=f'{settime[1].text}') 
                cell = worksheet.cell(row=index, column=4, value=f'{Money}'.lstrip())
                cell = worksheet.cell(row=index, column=5, value=f'{PersonCount[1].text}')
                cell = worksheet.cell(row=index, column=6, value=f'{event}'.lstrip())
                cell = worksheet.cell(row=index, column=8, value=f'{place[0].text}'.lstrip())
                print(f'情報集める成功{index}')

            else:
                print(f'{element}-no--Information')
                column_url.append(f'{element}noInformation')
                cell = worksheet.cell(row=index, column=3, value='noInformation')

        else:
            print(f'Failed to fetch: {{element}}')
    except requests.exceptions.RequestException as e:
        print(f'この会社はdodaに登録していない')
    # 保存 Excel 文件
    workbook.save('result.xlsx')

    # 关闭 Excel 文件
    workbook.close()
    # 指定要删除的文件路径
    file_path = "example.txt"

try:
    # 尝试删除文件
    os.remove('url.xlsx')
    print(f"{'url.xlsx'} 已删除")
except FileNotFoundError:
    # 如果文件不存在，打印提示信息
    print(f"{'url.xlsx'} 不存在")
except Exception as e:
    # 处理其他异常情况
    print(f"删除 {'url.xlsx'} 时发生错误: {e}")
try:
    # 尝试删除文件
    os.remove('CompanyName.xlsx')
    print(f"{'CompanyName.xlsx'} 已删除")
except FileNotFoundError:
    # 如果文件不存在，打印提示信息
    print(f"{'CompanyName.xlsx'} 不存在")
except Exception as e:
    # 处理其他异常情况
    print(f"删除 {'CompanyName.xlsx'} 时发生错误: {e}")
print('プログラム終わり、結果はresult.xlsx')